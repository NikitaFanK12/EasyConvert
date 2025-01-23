from tkinter import Tk, Label, Button, Entry, Canvas, Toplevel, OptionMenu, StringVar, filedialog, messagebox, \
    simpledialog, ttk
from docx2pdf import convert as docx_pdf_convert
from pdf2docx import Converter as pdf_docx_convert
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas as ReportLabCanvas
import pdfplumber
from openpyxl import Workbook
from PIL import Image, ImageTk
from pydub import AudioSegment
from moviepy.editor import VideoFileClip
import subprocess
import os
import sys


def convert_file(file_path=None, output_folder=None, inp_format='JPG', out_format='PNG', new_file_name="__New_File__", ffmpeg_path=None, ffprobe_path=None):
    if not file_path:
        return "Будь ласка, виберіть файл для конвертації."
    if not output_folder:
        return "Будь ласка, виберіть папку для збереження."

    if not ffmpeg_path:
        return "Для роботи бібліотеки easyconvert потрібно вказати шлях до ffmpeg."
    if not ffprobe_path:
        return "Для роботи бібліотеки easyconvert потрібно вказати шлях до ffprobe."

    input_format = inp_format
    output_format = out_format

    if not input_format or not output_format:
        return "Будь ласка, виберіть вхідний та вихідний формати."

    if new_file_name != "__New_File__":
        output_name = new_file_name
    else:
        output_name = ''

    if not output_name:
        return "Назва файлу не може бути порожньою."


    if output_format == "MP4-H265":
        output_file = os.path.join(output_folder, output_name + '.' + "mp4")
    else:
        output_file = os.path.join(output_folder, output_name + '.' + output_format.lower())

    try:
        if input_format in ['JPG', 'PNG', 'WEBP']:
            img = Image.open(file_path)
            if input_format in ['JPG', 'PNG'] and output_format == 'ICO':
                img.save(output_file, format='ICO')  # Збереження як ICO
            elif img.mode == 'P':  # Перетворюємо палітровий режим в RGB
                img = img.convert('RGB')
            elif input_format == 'WEBP' and output_format == 'PNG':
                img.save(output_file, format='PNG')
            img.save(output_file)

        elif input_format == 'CUR' and output_format == 'PNG':
            img = Image.open(file_path)
            img.save(output_file, format='PNG')

        elif input_format == 'DOCX' and output_format == 'PDF':
            docx_pdf_convert(file_path, output_file)

        elif input_format == 'PDF':
            if output_format == 'DOCX':
                cv = pdf_docx_convert(file_path)
                cv.convert(output_file, start=0, end=None)  # Конвертуємо весь файл
                cv.close()
            elif output_format == 'XLSX':
                workbook = Workbook()
                sheet = workbook.active
                with pdfplumber.open(file_path) as pdf:
                    for page in pdf.pages:
                        text = page.extract_text().splitlines()
                        for line in text:
                            sheet.append(line.split())  # Розбиваємо рядок на клітинки
                workbook.save(output_file)

        elif input_format == 'XLSX' and output_format == 'PDF':
            # Відкриваємо файл XLSX
            workbook = load_workbook(file_path)
            sheet = workbook.active

            # Налаштування PDF
            c = ReportLabCanvas.Canvas(output_file, pagesize=letter)
            width, height = letter
            y = height - 40

            for row in sheet.iter_rows(values_only=True):
                text = " ".join([str(cell) if cell is not None else "" for cell in row])

                c.drawString(30, y, text)
                y -= 20

                if y < 40:
                    c.showPage()
                    y = height - 40

            # Зберігаємо PDF
            c.save()

        elif input_format in ['M4A', 'MP3', 'WAV']:
            audio = AudioSegment.from_file(file_path, format=input_format.lower())
            audio.export(output_file, format=output_format.lower())

        elif input_format == 'MP4':
            video = VideoFileClip(file_path)
            if output_format == 'AVI':
                video.write_videofile(output_file, codec='png')  # Конвертуємо в AVI
            elif output_format == 'MP3':
                video.audio.write_audiofile(output_file)  # Витягуємо аудіо з відео та конвертуємо в MP3
            elif output_format == 'MP4-H265':  # Конвертуємо в MP4 за допомогою H.265
                video.write_videofile(output_file, codec='libx265')  # Використовуємо кодек H.265

        elif input_format == 'AVI' and output_format == 'MP4':
            video = VideoFileClip(file_path)
            video.write_videofile(output_file, codec='libx264')

        elif input_format == 'MKV' and output_format == 'MP4':
            video = VideoFileClip(file_path)
            video.write_videofile(output_file, codec='libx264')

        return 'true'
    except Exception as e:
        return "Не вдалося конвертувати файл: {e}"
