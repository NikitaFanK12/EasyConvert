from tkinter import Tk, Label, Button, Entry, Canvas, Toplevel, OptionMenu, StringVar, filedialog, messagebox, \
    simpledialog, ttk
from docx2pdf import convert as docx_pdf_convert
from pdf2docx import Converter as pdf_docx_convert
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas as ReportLabCanvas
import pytube
from pytube import YouTube
import pdfplumber
from openpyxl import Workbook
from PIL import Image, ImageTk
from pydub import AudioSegment
from moviepy.editor import VideoFileClip
import subprocess
import os
import sys

# Доступні конвертаціїї
conversions = {
    'JPG': ['PNG', 'ICO'],
    'PNG': ['JPG', 'ICO', 'CUR'],
    'M4A': ['MP3', 'WAV', 'OGG'],
    'MP3': ['WAV', 'OGG', 'M4A'],
    'WAV': ['MP3', 'OGG', 'M4A'],
    'CUR': ['PNG'],
    'DOCX': ['PDF'],
    'PDF': ['DOCX', 'XLSX'],
    'XLSX': ["PDF"],
    'MP4': ['AVI', 'MP3', 'MP4-H265'],
    'AVI': ['MP4'],
    'MKV': ['MP4'],
    'WEBP': ['PNG']
}

def get_ffmpeg_path():
    """Отримує шлях до ffmpeg.exe, доданого до пакету .exe"""
    if getattr(sys, 'frozen', False):
        ffmpeg_path = os.path.join(sys._MEIPASS, 'ffmpeg.exe')
        ffprobe_path = os.path.join(sys._MEIPASS, 'ffprobe.exe')  # Додаємо ffprobe
    else:
        ffmpeg_path = os.path.join(os.getcwd(), 'ffmpeg.exe')
        ffprobe_path = os.path.join(os.getcwd(), 'ffprobe.exe')

    if os.path.exists(ffmpeg_path) and os.path.exists(ffprobe_path):
        return ffmpeg_path, ffprobe_path
    else:
        raise FileNotFoundError("Не вдалося знайти ffmpeg.exe або ffprobe.exe")

def get_path(file_path):
    """Отримує шлях до file_path, доданого до пакету .exe"""
    if getattr(sys, 'frozen', False):
        path = os.path.join(sys._MEIPASS, file_path)
        path = os.path.join(sys._MEIPASS, file_path)
    else:
        path = os.path.join(os.getcwd(), file_path)
        path = os.path.join(os.getcwd(), file_path)

    if os.path.exists(path):
        return path
    else:
        raise FileNotFoundError("Не вдалося знайти file_path")


# Налаштовуємо шлях до ffmpeg.exe та ffprobe.exe
ffmpeg_path, ffprobe_path = get_ffmpeg_path()
AudioSegment.converter = ffmpeg_path

# Функція для конвертації
def convert_file(custom_format=False, inp_format='JPG', out_format='PNG', new_file_name="New File"):
    if not file_path:
        messagebox.showwarning("Помилка", "Будь ласка, виберіть файл для конвертації.")
        return
    if not output_folder:
        messagebox.showwarning("Помилка", "Будь ласка, виберіть папку для збереження.")
        return

    if custom_format:
        input_format = inp_format
        output_format = out_format
    else:
        input_format = input_var.get()
        output_format = output_var.get()

    if not input_format or not output_format:
        messagebox.showwarning("Помилка", "Будь ласка, виберіть вхідний та вихідний формати.")
        return

    # Запитуємо назву для нового файлу
    if new_file_name != "New File":
        output_name = new_file_name
    else:
        output_name = simpledialog.askstring("Назва файлу", "Введіть назву для конвертованого файлу (без розширення):")
    if not output_name:
        messagebox.showwarning("Помилка", "Назва файлу не може бути порожньою.")
        return

    # Формуємо шлях до вихідного файлу в обраній папці
    if output_format == "MP4-H265":
        output_file = os.path.join(output_folder, output_name + '.' + "mp4")
    else:
        output_file = os.path.join(output_folder, output_name + '.' + output_format.lower())

    try:
        if input_format in ['JPG', 'PNG', 'WEBP']:
            img = Image.open(file_path)
            if input_format in ['JPG', 'PNG'] and output_format == 'ICO':
                img.save(output_file, format='ICO')  # Збереження як ICO
            elif img.mode == 'P':  # Перетворюємо палітровий режим в RGB
                img = img.convert('RGB')
            elif input_format == 'WEBP' and output_format == 'PNG':
                img.save(output_file, format='PNG')
            img.save(output_file)

        elif input_format == 'CUR' and output_format == 'PNG':
            img = Image.open(file_path)
            img.save(output_file, format='PNG')

        elif input_format == 'DOCX' and output_format == 'PDF':
            docx_pdf_convert(file_path, output_file)

        elif input_format == 'PDF':
            if output_format == 'DOCX':
                cv = pdf_docx_convert(file_path)
                cv.convert(output_file, start=0, end=None)  # Конвертуємо весь файл
                cv.close()
            elif output_format == 'XLSX':
                workbook = Workbook()
                sheet = workbook.active
                with pdfplumber.open(file_path) as pdf:
                    for page in pdf.pages:
                        text = page.extract_text().splitlines()
                        for line in text:
                            sheet.append(line.split())  # Розбиваємо рядок на клітинки
                workbook.save(output_file)

        elif input_format == 'XLSX' and output_format == 'PDF':
            # Відкриваємо файл XLSX
            workbook = load_workbook(file_path)
            sheet = workbook.active

            # Налаштування PDF
            c = ReportLabCanvas.Canvas(output_file, pagesize=letter)
            width, height = letter
            y = height - 40

            for row in sheet.iter_rows(values_only=True):
                text = " ".join([str(cell) if cell is not None else "" for cell in row])

                c.drawString(30, y, text)
                y -= 20

                if y < 40:
                    c.showPage()
                    y = height - 40

            # Зберігаємо PDF
            c.save()

        elif input_format in ['M4A', 'MP3', 'WAV']:
            audio = AudioSegment.from_file(file_path, format=input_format.lower())
            audio.export(output_file, format=output_format.lower())

        elif input_format == 'MP4':
            video = VideoFileClip(file_path)
            if output_format == 'AVI':
                video.write_videofile(output_file, codec='png')  # Конвертуємо в AVI
            elif output_format == 'MP3':
                video.audio.write_audiofile(output_file)  # Витягуємо аудіо з відео та конвертуємо в MP3
            elif output_format == 'MP4-H265':  # Конвертуємо в MP4 за допомогою H.265
                video.write_videofile(output_file, codec='libx265')  # Використовуємо кодек H.265

        elif input_format == 'AVI' and output_format == 'MP4':
            video = VideoFileClip(file_path)
            video.write_videofile(output_file, codec='libx264')

        elif input_format == 'MKV' and output_format == 'MP4':
            create_progress_bar()
            video = VideoFileClip(file_path)
            video.write_videofile(output_file, codec='libx264')
            end_progress_bar()

        # Оновлюємо текст вихідного файлу
        output_file_label.config(text=f"Вихідний файл: {output_file}")
        messagebox.showinfo("Готово", f"Файл збережено як {output_file}")
    except Exception as e:
        messagebox.showerror("Помилка", f"Не вдалося конвертувати файл: {e}")


# Оновлення списку доступних форматів
def update_output_options(*args):
    output_var.set("")
    output_menu['menu'].delete(0, 'end')
    for option in conversions.get(input_var.get(), []):
        output_menu['menu'].add_command(label=option, command=lambda val=option: output_var.set(val))


# Вибір файлу
def select_file():
    global file_path
    file_path = filedialog.askopenfilename()
    if file_path:
        input_file_label.config(text=f"Вхідний файл: {file_path}")
        messagebox.showinfo("Файл обрано", f"Обрано файл: {file_path}")


# Вибір папки для збереження
def select_folder():
    global output_folder
    output_folder = filedialog.askdirectory()
    if output_folder:
        messagebox.showinfo("Папка обрана", f"Файли будуть збережені в: {output_folder}")

# Функція для створення прогрес-бару
def create_progress_bar():
    # global progress_bar, pb
    # pb = Toplevel(root)
    # pb.geometry("400x80")
    # pb.title("EasyConvert - Конвертація..")
    # progress_bar = ttk.Progressbar(pb, orient="horizontal", length=380, mode="determinate")
    # progress_bar.pack(pady=20)
    # progress_bar['value'] = 24
    # pb.update_idletasks()
    messagebox.showinfo("EasyConvert", "Почалася конвертація, натисніть ОК та чекайте кінця.")

def end_progress_bar():
    pass
#     for i in range(76):
#         root.after(10, add_progress)
#         pb.update_idletasks()
#     pb.destroy()

# def add_progress():
#     progress_bar['value'] += 1

def download_youtube_video(url, out_folder):
    try:
        # subprocess.run([
        #     "yt-dlp",
        #     "-o", f"{output_folder}",#/%(title)s.%(ext)s",
        #     url
        # ], check=True)
        #subprocess.run('yt-dlp -o "C://Users//Нікіта//Videos//video"  https://www.youtube.com/watch?v=7qQmEsvnDow -k -f mp4')
        yt_dlp_path = get_path('yt-dlp.exe')
        proc = subprocess.Popen(f'{yt_dlp_path} -o "{out_folder}//video" {url}')
        proc.wait()
        print("Відео завантажено успішно.")
    except subprocess.CalledProcessError as e:
        print(f"Помилка при завантаженні: {e}")

video_url = "https://www.youtube.com/watch?v=7qQmEsvnDow"
output_folder = "C://Users//Нікіта//Videos"
#download_youtube_video(video_url, output_folder)

# Test url:   https://www.youtube.com/watch?v=HpRSg1sVSYc
def open_youtube_window():
    def on_download_btn_click():
        inp = inp_url.get()
        download_youtube_video(url=inp, out_folder=output_folder)

    def on_download_music_btn_click():
        global file_path
        inp = inp_url.get()
        download_youtube_video(url=inp, out_folder=output_folder)
        file_path = f"{output_folder}//video.mkv"
        convert_file(True, 'MKV', 'MP4', "downloaded_video")
        file_path = f"{output_folder}//downloaded_video.mp4"
        convert_file(True, 'MP4', 'MP3', "video_music")

    tplvl = Toplevel(root)
    tplvl.title("EasyConvert - Завантаження відео з Youtube")
    tplvl.geometry(f"{WIDTH}x{HEIGHT}")

    canvas_yt = Canvas(tplvl, width=WIDTH, height=HEIGHT)
    canvas_yt.pack(fill="both", expand=True)
    canvas_yt.create_image(0, 0, anchor="nw", image=bg_image)
    canvas_yt.image = bg_image

    lb_url = Label(tplvl, text="Будьласка, введіть посилання (URL) до відео з YouTube:", bg=btn_color, fg=text_color)
    canvas_yt.create_window(WIDTH/2, 80, window=lb_url)

    lb_warn = Label(tplvl, text="Наразі ця функція на стадії розробки та може працювати не коректно (деякі посилання не будуть завантажуватися)!", bg=btn_color, fg=text_color)
    canvas_yt.create_window(WIDTH/2, 140, window=lb_warn)

    inp_url = Entry(tplvl, width=80, bg=btn_color, fg=text_color)
    canvas_yt.create_window(WIDTH/2, 100, window=inp_url)

    btn_download_yt_video = Button(tplvl, text="Завантажити відео з Youtube", bg=btn_color, fg=text_color, command=on_download_btn_click)
    btn_download_yt_video_audio = Button(tplvl, text="Завантажити аудіо з відео на Youtube", bg=btn_color, fg=text_color, command=on_download_music_btn_click)
    btn_select_folder_to_download = Button(tplvl, text="Вибрати папку де збережеться файл", bg=btn_color, fg=text_color
                                           , command=select_folder)
    canvas_yt.create_window(100, (HEIGHT - 20), window=btn_download_yt_video)
    canvas_yt.create_window(300, (HEIGHT - 20), window=btn_download_yt_video_audio)
    canvas_yt.create_window(120, (HEIGHT - 50), window=btn_select_folder_to_download)

    close_button = Button(tplvl, text="Назад", command=tplvl.destroy, bg=btn_color, fg=text_color)
    canvas_yt.create_window((WIDTH - 25), 15, window=close_button)

def info_menu():
    app_name = "EasyConvert"
    version = "1.0.23"
    author = "Nikita_K12"
    used_appslibs = "ffmpeg, yt-dlp"

    infmenu = Toplevel(root)
    infmenu.geometry("620x320")
    infmenu.title(f"{app_name} - Інформація")
    infmenu.config(bg=btn_color)

    text = (f"{app_name}\n\n"
            f"Product name: {app_name}\n"
            f"Version: {version}\n"
            f"Author: {author}\n\n"
            f"Thanks to: {used_appslibs}\n")
    lb = Label(infmenu, text=text, bg=btn_color, fg=text_color)
    lb.place(x=0, y=0)

    help_text = ("Допомога\n"
                 "Вітаємо вас у програмі EasyConvert! Тут ви зможете перекон-\n"
                 "вертувати ваші файли, та навіть завантажувати відео та аудіо з\n"
                 "YouTube. Розпочнімо. Якщо ви хочете переконвертувати ваш вайл\n"
                 "формату JPG в файл формату PNG, то спочатку виберіть вхідний\n"
                 "формат (в нашому випадку JPG). Далі виберіть вихідний формат.\n"
                 "Це той формат файлу який ми хочемо отримати. Після того як ми\n"
                 "вибрали необіхдні формати, треба вибрати сам файл який ми\n"
                 "хочемо переконвертувати (наприклад my_photo.jpg). Для цього\n"
                 "натисність кнопку 'Обрати файл для конвертації', та оберіть\n"
                 "файл. Далі треба вибрати папку куди збережеться переконвер-\n"
                 "тований файл, для цього натисніть кнопку 'Обрати папку де\n"
                 "збережеться файл'. Після того як ми все обрали, нам залишилося\n"
                 "натиснути кнопку 'Конвертувати', ввести назву нового файлу, та\n"
                 "чекати завершення. Коли з'явиться вікно 'Конвертація завершена',\n"
                 "ми зможемо побачити переконвертований файл в папці яку ми зазначили.")
    help_lb = Label(infmenu, text=help_text,bg=btn_color, fg=text_color)
    help_lb.place(x=180, y=0)


# Ініціалізація графічного інтерфейсу
root = Tk()
root.title("EasyConvert - Конвертер файлів")

WIDTH = 600
HEIGHT = 320
btn_color = "#d6ac79"
text_color = "#422c10"

root.geometry(f"{WIDTH}x{HEIGHT}")

canvas = Canvas(root, width=WIDTH, height=HEIGHT)
canvas.pack(fill="both", expand=True)

bg_image = ImageTk.PhotoImage(Image.open(get_path("bg.png")))
canvas.create_image(0, 0, anchor="nw", image=bg_image)
canvas.image = bg_image

file_path = ""  # Змінна для збереження шляху до файлу
output_folder = ""  # Змінна для збереження шляху до папки для збереження

app_name_lb = Label(root, text="EasyConvert", font=("Arial", 11, "bold"), bg=btn_color, fg=text_color)
canvas.create_window(50, 14, window=app_name_lb)
app_info_btn = Button(root, text="Інформація", bg=btn_color, fg=text_color, command=info_menu)
canvas.create_window(45, 40, window=app_info_btn, width=80, height=22)

# Віджети для показу вибраного файлу та вихідного файлу
input_file_label = Label(root, text="Вхідний файл: не вибрано", bg=btn_color, fg=text_color)
canvas.create_window(WIDTH/2, 15, window=input_file_label)
output_file_label = Label(root, text="Вихідний файл: не визначено", bg=btn_color, fg=text_color)
canvas.create_window(WIDTH/2, 45, window=output_file_label)

lb_in_format = Label(root, text="Виберіть вхідний формат:", bg=btn_color, fg=text_color)
lb_out_format = Label(root, text="Виберіть вихідний формат:", bg=btn_color, fg=text_color)
canvas.create_window(75, 100, window=lb_in_format)
canvas.create_window(375, 100, window=lb_out_format)

input_var = StringVar(root)
input_menu = OptionMenu(root, input_var, *conversions.keys())
canvas.create_window(75, 130, window=input_menu)
input_var.trace("w", update_output_options)

output_var = StringVar(root)
output_menu = OptionMenu(root, output_var, '')
canvas.create_window(375, 130, window=output_menu)

btn_select_file = Button(root, text="Обрати файл для конвертації", bg=btn_color, fg=text_color, command=select_file)
btn_select_folder = Button(root, text="Вибрати папку де збережеться файл", bg=btn_color, fg=text_color, command=select_folder)
btn_convert = Button(root, text="Конвертувати", bg=btn_color, fg=text_color, command=convert_file)
btn_yt_download_menu = Button(root, text="Завантажити відео/аудіо з Youtube", bg=btn_color, fg=text_color, command=open_youtube_window)
canvas.create_window(89, (HEIGHT - 20), window=btn_select_file)
canvas.create_window(285, (HEIGHT - 20), window=btn_select_folder)
canvas.create_window(104, (HEIGHT - 50), window=btn_yt_download_menu)
canvas.create_window((WIDTH - 45), (HEIGHT - 15), window=btn_convert)

root.mainloop()
